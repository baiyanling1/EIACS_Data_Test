import csv
import random
import time
import uuid
from random import randint
import openpyxl
import datetime
import pandas as pd
from id_num import generate_id_number

now = datetime.datetime.now()
formatted_time = now.strftime('%Y%m%d%H%M%S')

surname = ["老师","学生","校长","董事","家属"]
IP_group = ["192.168.0.1/24","192.168.0.1/16","192.168.0.1/32","192.168.0.1/8"]

create_at = time.strftime('%Y-%m-%d %H:%M:%S')
update_at = time.strftime('%Y-%m-%d %H:%M:%S')

str_phone = list()
str_id = list()
str_id_connt = list()
str_phone_connt = list()



def randomnum():  # 手机号
    str_phone.append(str(15))
    for i in range(0, 9):
        num = randint(1, 9)
        str_phone.append(str(num))
    if str_phone not in str_phone_connt:
        # print(str_phone)
        return str_phone
    else:
        pass


def deal_day(d):  # 日期处理
    id_day = randint(1, d)
    if id_day < 10:
        str_id.append("0")
        str_id.append(str(id_day))
    else:
        str_id.append(str(id_day))


def randid():  # 身份证
    list1 = ['11', '12', '13', '14', '15', '21', '22', '23', '31', '32', '33', '34', '35', '36', '37', '41', '42', '43',
             '44', '45', '46', '50', '51', '52', '53', '54', '61', '62', '63', '64', '65', '81', '82', '83']  # 全国区域代码
    province_num = randint(1, len(list1) - 1)
    str_id.append(list1[province_num])
    city_num = randint(1000, 9999)
    str_id.append(str(city_num))
    id_year = randint(1950, 2010)
    str_id.append(str(id_year))
    id_month = randint(1, 12)
    if id_month < 10:
        str_id.append("0")
        str_id.append(str(id_month))
    else:
        str_id.append(str(id_month))
    if id_year % 4 == 0:  # 闰年
        if id_month == 2:
            deal_day(29)
        elif id_month in (1, 3, 5, 7, 8, 10, 12):
            deal_day(31)
        else:
            deal_day(30)
    else:
        if id_month == 2:
            deal_day(28)
        elif id_month in (1, 3, 5, 7, 8, 10, 12):
            deal_day(31)
        else:
            deal_day(30)
    rand_num = randint(100, 999)
    str_id.append(str(rand_num))
    check_num = randint(1, 10)
    if check_num == 10:
        str_id.append("x")
    else:
        str_id.append(str(check_num))
    # print(str_id)
    return str_id


def randcard():  # 银行卡号
    str_card = list()
    str_card.append(str(621799126))
    rand_card_num1 = str(randint(10000, 49999))
    str_card.append(rand_card_num1)
    rand_card_num2 = str(randint(50000, 99999))
    str_card.append(rand_card_num2)
    return str_card


def randname():  # 姓名
    xing = [
        '赵', '钱', '孙', '李', '周', '吴', '郑', '王', '冯', '陈', '褚', '卫', '蒋', '沈', '韩', '杨', '朱', '秦',
        '尤', '许',
        '何', '吕', '施', '张', '孔', '曹', '严', '华', '金', '魏', '陶', '姜', '戚', '谢', '邹', '喻', '柏', '水',
        '窦', '章',
        '云', '苏', '潘', '葛', '奚', '范', '彭', '郎', '鲁', '韦', '昌', '马', '苗', '凤', '花', '方', '俞', '任',
        '袁', '柳',
        '酆', '鲍', '史', '唐', '费', '廉', '岑', '薛', '雷', '贺', '倪', '汤', '滕', '殷', '罗', '毕', '郝', '邬',
        '安', '常',
        '乐', '于', '时', '傅', '皮', '卞', '齐', '康', '伍', '余', '元', '卜', '顾', '孟', '平', '黄', '和', '穆',
        '萧', '尹',
        '姚', '邵', '堪', '汪', '祁', '毛', '禹', '狄', '贝', '明', '臧', '计', '伏', '成', '戴', '谈', '宋', '茅',
        '庞', '梁']
    ming1 = ['一', '二', '三', '四', '五', '六', '七', '八', '九']
    ming2 = [
        '的', '一', '是', '了', '我', '不', '人', '在', '他', '有', '这', '个', '上', '们', '来', '到', '时', '大',
        '地', '为',
        '子', '中', '你', '说', '生', '国', '年', '着', '就', '那', '和', '要', '她', '出', '也', '得', '里', '后',
        '自', '以',
        '会', '家', '可', '下', '而', '过', '天', '去', '能', '对', '小', '多', '然', '于', '心', '学', '么', '之',
        '都', '好',
        '看', '起', '发', '当', '没', '成', '只', '如', '事', '把', '还', '用', '第', '样', '道', '想', '作', '种',
        '开', '美',
        '总', '从', '无', '情', '己', '面', '最', '女', '但', '现', '前', '些', '所', '同', '日', '手', '又', '行',
        '意', '动',
        '方', '期', '它', '头', '经', '长', '儿', '回', '位', '分', '爱', '老', '因', '很', '给', '名', '法', '间',
        '斯', '知',
        '世', '什', '两', '次', '使', '身', '者', '被', '高', '已', '亲', '其', '进', '此', '话', '常', '与', '活',
        '正', '感',
        '见', '明', '问', '力', '理', '尔', '点', '文', '几', '定', '本', '公', '特', '做', '外', '孩', '相', '西',
        '果', '走',
        '将', '月', '十', '实', '向', '声', '车', '全', '信', '重', '三', '机', '工', '物', '气', '每', '并', '别',
        '真', '打',
        '太', '新', '比', '才', '便', '夫', '再', '书', '部', '水', '像', '眼', '等', '体', '却', '加', '电', '主',
        '界', '门',
        '利', '海', '受', '听', '表', '德', '少', '克', '代', '员', '许', '稜', '先', '口', '由', '死', '安', '写',
        '性', '马',
        '光', '白', '或', '住', '难', '望', '教', '命', '花', '结', '乐', '色', '更', '拉', '东', '神', '记', '处',
        '让', '母',
        '父', '应', '直', '字', '场', '平', '报', '友', '关', '放', '至', '张', '认', '接', '告', '入', '笑', '内',
        '英', '军',
        '候', '民', '岁', '往', '何', '度', '山', '觉', '路', '带', '万', '男', '边', '风', '解', '叫', '任', '金',
        '快', '原',
        '吃', '妈', '变', '通', '师', '立', '象', '数', '四', '失', '满', '战', '远', '格', '士', '音', '轻', '目',
        '条', '呢',
        '病', '始', '达', '深', '完', '今', '提', '求', '清', '王', '化', '空', '业', '思', '切', '怎', '非', '找',
        '片', '罗',
        '钱', '紶', '吗', '语', '元', '喜', '曾', '离', '飞', '科', '言', '干', '流', '欢', '约', '各', '即', '指',
        '合', '反',
        '题', '必', '该', '论', '交', '终', '林', '请', '医', '晚', '制', '球', '决', '窢', '传', '画', '保', '读',
        '运', '及',
        '则', '房', '早', '院', '量', '苦', '火', '布', '品', '近', '坐', '产', '答', '星', '精', '视', '五', '连',
        '司', '巴',
        '奇', '管', '类', '未', '朋', '且', '婚', '台', '夜', '青', '北', '队', '久', '乎', '越', '观', '落', '尽',
        '形', '影',
        '红', '爸', '百', '令', '周', '吧', '识', '步', '希', '亚', '术', '留', '市', '半', '热', '送', '兴', '造',
        '谈', '容',
        '极', '随', '演', '收', '首', '根', '讲', '整', '式', '取', '照', '办', '强', '石', '古', '华', '諣', '拿',
        '计', '您',
        '装', '似', '足', '双', '妻', '尼', '转', '诉', '米', '称', '丽', '客', '南', '领', '节', '衣', '站', '黑',
        '刻', '统',
        '断', '福', '城', '故', '历', '惊', '脸', '选', '包', '紧', '争', '另', '建', '维', '绝', '树', '系', '伤',
        '示', '愿',
        '持', '千', '史', '谁', '准', '联', '妇', '纪', '基', '买', '志', '静', '阿', '诗', '独', '复', '痛', '消',
        '社', '算',
        '义', '竟', '确', '酒', '需', '单', '治', '卡', '幸', '兰', '念', '举', '仅', '钟', '怕', '共', '毛', '句',
        '息', '功',
        '官', '待', '究', '跟', '穿', '室', '易', '游', '程', '号', '居', '考', '突', '皮', '哪', '费', '倒', '价',
        '图', '具',
        '刚', '脑', '永', '歌', '响', '商', '礼', '细', '专', '黄', '块', '脚', '味', '灵', '改', '据', '般', '破',
        '引', '食',
        '仍', '存', '众', '注', '笔', '甚', '某', '沉', '血', '备', '习', '校', '默', '务', '土', '微', '娘', '须',
        '试', '怀',
        '料', '调', '广', '蜖', '苏', '显', '赛', '查', '密', '议', '底', '列', '富', '梦', '错', '座', '参', '八',
        '除', '跑',
        '亮', '假', '印', '设', '线', '温', '虽', '掉', '京', '初', '养', '香', '停', '际', '致', '阳', '纸', '李',
        '纳', '验',
        '助', '激', '够', '严', '证', '帝', '饭', '忘', '趣', '支', '春', '集', '丈', '木', '研', '班', '普', '导',
        '顿', '睡',
        '展', '跳', '获', '艺', '六', '波', '察', '群', '皇', '段', '急', '庭', '创', '区', '奥', '器', '谢', '弟',
        '店', '否',
        '害', '草', '排', '背', '止', '组', '州', '朝', '封', '睛', '板', '角', '况', '曲', '馆', '育', '忙', '质',
        '河', '续',
        '哥', '呼', '若', '推', '境', '遇', '雨', '标', '姐', '充', '围', '案', '伦', '护', '冷', '警', '贝', '著',
        '雪', '索',
        '剧', '啊', '船', '险', '烟', '依', '斗', '值', '帮', '汉', '慢', '佛', '肯', '闻', '唱', '沙', '局', '伯',
        '族', '低',
        '玩', '资', '屋', '击', '速', '顾', '泪', '洲', '团', '圣', '旁', '堂', '兵', '七', '露', '园', '牛', '哭',
        '旅', '街',
        '劳', '型', '烈', '姑', '陈', '莫', '鱼', '异', '抱', '宝', '权', '鲁', '简', '态', '级', '票', '怪', '寻',
        '杀', '律',
        '胜', '份', '汽', '右', '洋', '范', '床', '舞', '秘', '午', '登', '楼', '贵', '吸', '责', '例', '追', '较',
        '职', '属',
        '渐', '左', '录', '丝', '牙', '党', '继', '托', '赶', '章', '智', '冲', '叶', '胡', '吉', '卖', '坚', '喝',
        '肉', '遗',
        '救', '修', '松', '临', '藏', '担', '戏', '善', '卫', '药', '悲', '敢', '靠', '伊', '村', '戴', '词', '森',
        '耳', '差',
        '短', '祖', '云', '规', '窗', '散', '迷', '油', '旧', '适', '乡', '架', '恩', '投', '弹', '铁', '博', '雷',
        '府', '压',
        '超', '负', '勒', '杂', '醒', '洗', '采', '毫', '嘴', '毕', '九', '冰', '既', '状', '乱', '景', '席', '珍',
        '童', '顶',
        '派', '素', '脱', '农', '疑', '练', '野', '按', '犯', '拍', '征', '坏', '骨', '余', '承', '置', '臓', '彩',
        '灯', '巨',
        '琴', '免', '环', '姆', '暗', '换', '技', '翻', '束', '增', '忍', '餐', '洛', '塞', '缺', '忆', '判', '欧',
        '层', '付',
        '阵', '玛', '批', '岛', '项', '狗', '休', '懂', '武', '革', '良', '恶', '恋', '委', '拥', '娜', '妙', '探',
        '呀', '营',
        '退', '摇', '弄', '桌', '熟', '诺', '宣', '银', '势', '奖', '宫', '忽', '套', '康', '供', '优', '课', '鸟',
        '喊', '降',
        '夏', '困', '刘', '罪', '亡', '鞋', '健', '模', '败', '伴', '守', '挥', '鲜', '财', '孤', '枪', '禁', '恐',
        '伙', '杰',
        '迹', '妹', '藸', '遍', '盖', '副', '坦', '牌', '江', '顺', '秋', '萨', '菜', '划', '授', '归', '浪', '听',
        '凡', '预',
        '奶', '雄', '升', '碃', '编', '典', '袋', '莱', '含', '盛', '济', '蒙', '棋', '端', '腿', '招', '释', '介',
        '烧', '误',
        '乾', '坤']
    x = randint(0, len(xing) - 1)
    y = randint(0, len(ming1) - 1)
    z = randint(0, len(ming2) - 1)
    # print(xing[x] + ming1[y] + ming2[z])
    return xing[x] + ming1[y] + ming2[z]


def show_info():  # 显示函数
    print("姓    名：%s" % ("".join(randname())))
    print("手 机 号：%s" % ("".join(randomnum())))
    print("身份证号：%s" % ("".join(randid())))
    print("银行卡号：%s" % ("".join(randcard())))

def get_auth_user(num):
    wb = openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.title = "参数说明"
    sheet2 = wb.create_sheet("Sheet2")
    sheet2.title = "鉴权终端用户数据"
    # 表头
    titles = (
    '姓名', '手机号*', '身份证', '分组', '部门', '认证方式（0免密认证，1密码认证）*', '认证用户名*', '认证密码', 'IP地址',
    'IP地址池名称', 'MAC地址', '备注')
    for index in range(len(titles)):
        # write的第一个参数：行，第二个参数：列 第三个参数：内容 第四个参数：样式
        sheet2.cell(row=1, column=index + 1).value = titles[index]
    x = ""
    y = ""
    z = ""
    # #写文件
    for r in range(1, num+1):
        for c in range(1, 13):
            if c == 1:
                sheet2.cell(r + 1, c).value = x.join(randname())
            if c == 2:
                sheet2.cell(r + 1, c).value = z.join(randomnum())
            if c == 3:
                sheet2.cell(r + 1, c).value = y.join(randid())
            if c == 4:
                sheet2.cell(r + 1, c).value = '普通用户组-TEST'
            # '部门',
            if c == 5:
                sheet2.cell(r + 1, c).value = '普通用户组-TEST'
            # '认证方式（0免密认证，1密码认证）*
            if c == 6:
                sheet2.cell(r + 1, c).value = '1'
            # 认证用户名*',
            if c == 7:
                sheet2.cell(r + 1, c).value = 'test' + str(r)
            # '证密码',
            if c == 8:
                sheet2.cell(r + 1, c).value = 'Redtea@123'
            # 'IP地址',
            if c == 9:
                sheet2.cell(r + 1, c).value = ''
            # 'IP地址池名称',
            if c == 10:
                sheet2.cell(r + 1, c).value = ''
            # 'MAC地址'
            if c == 11:
                sheet2.cell(r + 1, c).value = 'A0-10-10-B0-3A-88'
            # ,'备注'
            if c == 12:
                sheet2.cell(r + 1, c).value = '普通用户组-TEST' + str(r) + str(c)
            x = ""
            y = ""
            z = ""
            global str_phone
            global str_id
            str_phone = list()
            str_id = list()
    sheet2.column_dimensions['B'].width = 13
    sheet2.column_dimensions['C'].width = 20
    sheet2.column_dimensions['D'].width = 20
    sheet2.column_dimensions['E'].width = 20
    sheet2.column_dimensions['H'].width = 15
    sheet2.column_dimensions['K'].width = 20
    sheet2.column_dimensions['L'].width = 25
    wb.save(r'/Users/hejian/Desktop/联通数科/性能测试数据/鉴权性能测试数据/auth-user-'+str(num)+'-'+str(formatted_time)+'.xlsx')
   # 读取Excel文件
    df = pd.read_excel('/Users/hejian/Desktop/联通数科/性能测试数据/鉴权性能测试数据/auth-user-'+str(num)+'-'+str(formatted_time)+'.xlsx', sheet_name='鉴权终端用户数据')

    # 将数据保存为csv文件
    df.to_csv('/Users/hejian/Desktop/联通数科/性能测试数据/鉴权性能测试数据/auth-user-'+str(num)+'-'+str(formatted_time)+'.csv', index=False)

def get_auth_user_gpt(num):
    wb = openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.title = "参数说明"
    sheet2 = wb.create_sheet("鉴权终端用户数据")
    # 表头
    titles = ('姓名', '手机号*', '身份证', '分组', '部门', '认证方式（0免密认证，1密码认证）*', '认证用户名*', '认证密码', 'IP地址',
              'IP地址池名称', 'MAC地址', '备注')
    for index, title in enumerate(titles):
        sheet2.cell(row=1, column=index + 1, value=title)
    # 生成数据
    for r in range(2, num + 2):
        name = randname()
        phone = randomnum()
        id_num = randid()
        sheet2.cell(r, 1, value=name)
        sheet2.cell(r, 2, value="".join(phone))
        sheet2.cell(r, 3, value="".join(id_num))
        sheet2.cell(r, 4, value='普通用户组-TEST')
        sheet2.cell(r, 5, value='普通用户组-TEST')
        sheet2.cell(r, 6, value='1')
        sheet2.cell(r, 7, value='test' + str(r))
        sheet2.cell(r, 8, value='Redtea@123')
        sheet2.cell(r, 9, value='')
        sheet2.cell(r, 10, value='')
        sheet2.cell(r, 11, value='A0-10-10-B0-3A-88')
        sheet2.cell(r, 12, value='普通用户组-TEST' + str(r) + '12')
        global str_phone
        global str_id
        str_phone = list()
        str_id = list()
    # 设置列宽
    for col in ['B', 'C', 'D', 'E', 'H', 'K', 'L']:
        sheet2.column_dimensions[col].width = 15
    # 保存为Excel和CSV文件
    file_name = f'auth-user-{num}-{formatted_time}'
    wb.save(f'/Users/hejian/Desktop/联通数科/性能测试数据/鉴权性能测试数据/user/{file_name}.xlsx')
    df = pd.read_excel(f'/Users/hejian/Desktop/联通数科/性能测试数据/鉴权性能测试数据/user/{file_name}.xlsx', sheet_name='鉴权终端用户数据')
    df.to_csv(f'/Users/hejian/Desktop/联通数科/性能测试数据/鉴权性能测试数据/user/{file_name}.csv', index=False)
def get_auth_cpe_user(num):
    wb = openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.title = "参数说明"
    sheet2 = wb.create_sheet("Sheet2")
    sheet2.title = "鉴权CPE设备数据"

    # 表头
    titles = (
    '姓名', '手机号*', '身份证', '分组', '认证方式（0免密认证，1密码认证）*', '认证用户名*', '认证密码', 'IP地址',
    'IP地址池名称', 'MAC地址', '下挂IP地址段*','备注')
    for index in range(len(titles)):
        # write的第一个参数：行，第二个参数：列 第三个参数：内容 第四个参数：样式
        sheet2.cell(row=1, column=index + 1).value = titles[index]
    x = ""
    y = ""
    z = ""
    # #写文件
    for r in range(1, num+1):
        for c in range(1, 13):
            if c == 1:
                sheet2.cell(r + 1, c).value = x.join(randname())
            if c == 2:
                sheet2.cell(r + 1, c).value = z.join(randomnum())
            if c == 3:
                sheet2.cell(r + 1, c).value = y.join(randid())
            if c == 4:
                sheet2.cell(r + 1, c).value = 'CPE用户组-TEST'
            # '认证方式（0免密认证，1密码认证）*
            if c == 5:
                sheet2.cell(r + 1, c).value = '1'
            # 认证用户名*',
            if c == 6:
                sheet2.cell(r + 1, c).value = 'test' + str(r)
            # '证密码',
            if c == 7:
                sheet2.cell(r + 1, c).value = 'Redtea@123'
            # 'IP地址',
            if c == 8:
                sheet2.cell(r + 1, c).value = ''
            # 'IP地址池名称',
            if c == 9:
                sheet1.cell(r + 1, c).value = ''
            # 'MAC地址'
            if c == 10:
                sheet2.cell(r + 1, c).value = 'A0-10-10-B0-3A-88'
            # ,'备注'
            if c == 11:
                sheet2.cell(r + 1, c).value = str(random.choice(IP_group))
            if c == 12:
                sheet2.cell(r + 1, c).value = 'CPE用户组-TEST' + str(r) + str(c)
            x = ""
            y = ""
            z = ""
            global str_phone
            global str_id
            str_phone = list()
            str_id = list()
    sheet2.column_dimensions['B'].width = 13
    sheet2.column_dimensions['C'].width = 20
    sheet2.column_dimensions['D'].width = 20
    sheet2.column_dimensions['L'].width = 20
    sheet2.column_dimensions['G'].width = 15
    sheet2.column_dimensions['J'].width = 20
    sheet2.column_dimensions['K'].width = 25
    wb.save(r'/Users/hejian/Desktop/联通数科/性能测试数据/鉴权性能测试数据/auth-CPE-'+str(num)+'-'+str(formatted_time)+'.xlsx')
    # 读取Excel文件
    df = pd.read_excel('/Users/hejian/Desktop/联通数科/性能测试数据/鉴权性能测试数据/auth-CPE-'+str(num)+'-'+str(formatted_time)+'.xlsx', sheet_name='鉴权CPE设备数据')

    # 将数据保存为csv文件
    df.to_csv('/Users/hejian/Desktop/联通数科/性能测试数据/鉴权性能测试数据/auth-CPE-'+str(num)+'-'+str(formatted_time)+'.csv', index=False)

def get_auth_cpe_user_gpt(num):
    wb = openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.title = "参数说明"
    sheet2 = wb.create_sheet("鉴权CPE设备数据")

    # 表头
    titles = ('设备ID', '设备名称', '电话号码*', '分组', '认证方式（0免密认证，1密码认证）*', '认证用户名*', '认证密码', 'IP地址',
              'IP地址池名称', 'MAC地址', '下挂IP地址段*', '备注')
    for index, title in enumerate(titles):
        sheet2.cell(row=1, column=index + 1, value=title)
    # 生成数据
    for r in range(2, num + 2):
        name = randname()
        phone = randomnum()
        id_num = randid()
        sheet2.cell(r, 1, value=str(r))
        sheet2.cell(r, 2, value='CPE'+str(num))
        sheet2.cell(r, 3, value="".join(phone))
        sheet2.cell(r, 4, value='CPE用户组-TEST')
        sheet2.cell(r, 5, value='1')
        sheet2.cell(r, 6, value='CPE_test' + str(r))
        sheet2.cell(r, 7, value='Redtea@123')
        sheet2.cell(r, 8, value='')
        sheet2.cell(r, 9, value='')
        sheet2.cell(r, 10, value='A0-10-10-B0-3A-88')
        sheet2.cell(r, 11, value=random.choice(IP_group))
        sheet2.cell(r, 12, value='CPE用户组-TEST' + str(r) + '12')
        global str_phone
        global str_id
        str_phone = list()
        str_id = list()
    # 设置列宽
    for col in ['B', 'C', 'D', 'G', 'J', 'K', 'L']:
        sheet2.column_dimensions[col].width = 15
    sheet2.column_dimensions['L'].width = 20
    # 保存为Excel和CSV文件
    file_name = f'auth-CPE-{num}-{formatted_time}'
    wb.save(f'/Users/hejian/Desktop/联通数科/性能测试数据/鉴权性能测试数据/cpe/{file_name}.xlsx')
    df = pd.read_excel(f'/Users/hejian/Desktop/联通数科/性能测试数据/鉴权性能测试数据/cpe/{file_name}.xlsx', sheet_name='鉴权CPE设备数据')
    df.to_csv(f'/Users/hejian/Desktop/联通数科/性能测试数据/鉴权性能测试数据/cpe/{file_name}.csv', index=False)

if __name__ == '__main__':
    get_auth_cpe_user_gpt(100000)
    # get_auth_user_gpt(1000)