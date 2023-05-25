import base64
import csv
import random
import time
import uuid
from random import randint

import pandas as pd

from eiacs_data_base import DB_BIZ
import openpyxl
import  id_num

def encode_password(password):
    # 将字符串转换为二进制数据
    password_bytes = password.encode('utf-8')
    # 对二进制数据进行Base64编码
    encoded_password_bytes = base64.b64encode(password_bytes)
    # 将Base64编码后的二进制数据转换为字符串
    encoded_password = encoded_password_bytes.decode('utf-8')
    return encoded_password
#自己写的
def Insert_sign_user(filename):
    mysql_BIZ = DB_BIZ()
    mysql_BIZ.init()
    created_time = time.strftime('%Y-%m-%d %H:%M:%S')
    update_time = time.strftime('%Y-%m-%d %H:%M:%S')
    userValues_BIZ = []
    sql_BIZ_1 = "INSERT INTO user_signed_info(created_time, group_name, id_card, legal_user, msisdn,name,update_time) VALUE (%s,%s,%s,%s,%s,%s,%s)"

    workbook = openpyxl.load_workbook(filename)
    sheet = workbook['鉴权终端用户数据']
    # data = (psycopg2.Binary(b'1'),)
    # 遍历行
    for row in sheet.iter_rows(min_row=2):
        # 获取单元格数据
        name = row[0].value
        msisdn = row[1].value
        id_card = row[2].value
        group_name = row[3].value
        userValues_BIZ.append(
            (str(created_time), str(group_name), str(id_card), 1, str(msisdn), str(name), str(update_time)))
        if len(userValues_BIZ) == 1000:
            mysql_BIZ.insert_new_data(sql_BIZ_1, userValues_BIZ)
            userValues_BIZ = []

    mysql_BIZ.commit()

#gpt帮我优化的：将鉴权用户数据插入到已签约用户表中
def Insert_sign_user_new(filename):
    mysql_BIZ = DB_BIZ()
    mysql_BIZ.init()

    created_time = time.strftime('%Y-%m-%d %H:%M:%S')
    update_time = time.strftime('%Y-%m-%d %H:%M:%S')
    userValues_BIZ = []
    sql_BIZ_1 = "INSERT INTO user_signed_info(created_time, group_name, id_card, legal_user, msisdn,name,update_time) VALUES (%s,%s,%s,%s,%s,%s,%s)"

    workbook = openpyxl.load_workbook(filename)
    sheet = workbook['鉴权终端用户数据']

    # 遍历行
    for row in sheet.iter_rows(min_row=2):
        # 获取单元格数据
        name = row[0].value
        msisdn = row[1].value
        id_card = row[2].value
        group_name = row[3].value
        userValues_BIZ.append((created_time, group_name, id_card, 1, msisdn, name, update_time))

        # 如果userValues_BIZ列表长度达到1000，则批量插入到数据库中
        if len(userValues_BIZ) == 1000:
            mysql_BIZ.insert_new_data(sql_BIZ_1, userValues_BIZ)
            userValues_BIZ = []

    # 如果还有未插入的数据，则批量插入到数据库中
    if len(userValues_BIZ) > 0:
        mysql_BIZ.insert_new_data(sql_BIZ_1, userValues_BIZ)

    mysql_BIZ.commit()

#自己写的
def Delete_sign_user(filename):
    mysql_BIZ = DB_BIZ()
    mysql_BIZ.init()

    workbook = openpyxl.load_workbook(filename)
    sheet = workbook['鉴权终端用户数据']
    # 遍历行
    for row in sheet.iter_rows(min_row=2):
        # 获取单元格数据
        name = row[0].value
        msisdn = row[1].value
        id_card = row[2].value
        group_name = row[3].value
        mysql_BIZ.delete_data("DELETE FROM user_signed_info WHERE msisdn = "+str(msisdn))

    mysql_BIZ.commit()

#gpt帮我优化的
def Delete_sign_user_new(filename):
    mysql_BIZ = DB_BIZ()
    mysql_BIZ.init()

    workbook = openpyxl.load_workbook(filename)
    sheet = workbook['鉴权终端用户数据']
    # 获取所有用户数据的msisdn值，并拼接成一个逗号分隔的字符串
    msisdns = ",".join([str(row[1].value) for row in sheet.iter_rows(min_row=2)])

    # 执行单个SQL语句，批量删除所有用户数据
    mysql_BIZ.delete_data("DELETE FROM user_signed_info WHERE msisdn IN ({})".format(msisdns))

    mysql_BIZ.commit()
def Delete_common_user_new(filename):
    mysql_BIZ = DB_BIZ()
    mysql_BIZ.init()

    workbook = openpyxl.load_workbook(filename)
    sheet = workbook['企业用户数据']
    # 获取所有用户数据的msisdn值，并拼接成一个逗号分隔的字符串
    msisdns = ",".join([str(row[1].value) for row in sheet.iter_rows(min_row=2)])

    # 执行单个SQL语句，批量删除所有用户数据
    mysql_BIZ.delete_data("DELETE FROM biz_commons_user WHERE msisdn IN ({})".format(msisdns))

    mysql_BIZ.commit()

def Delete_common_user_auth_user(filename):
    mysql_BIZ = DB_BIZ()
    mysql_BIZ.init()

    workbook = openpyxl.load_workbook(filename)
    sheet = workbook['鉴权终端用户数据']
    # 获取所有用户数据的msisdn值，并拼接成一个逗号分隔的字符串
    msisdns = ",".join([str(row[1].value) for row in sheet.iter_rows(min_row=2)])

    # 执行单个SQL语句，批量删除所有用户数据
    mysql_BIZ.delete_data("DELETE FROM biz_commons_user WHERE msisdn IN ({})".format(msisdns))

    mysql_BIZ.commit()
def Insert_AuthData_to_mysql(filename):
    # 读取Excel文件
    mysql_BIZ = DB_BIZ()
    mysql_BIZ.init()

    created_time = time.strftime('%Y-%m-%d %H:%M:%S')
    update_time = time.strftime('%Y-%m-%d %H:%M:%S')
    userValues_BIZ = []
    sql_BIZ_1 = 'INSERT INTO db_eiacs_biz.biz_commons_user (active_num, auth_name, auth_pass, auth_type, create_time, department, group_name, id_card, ip, is_online, legal_user, mac_addr, msisdn, nas_id_str, offline_time, online_time, real_name, remark, update_time, session_expire_time, is_sign) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'

    workbook = openpyxl.load_workbook(filename)
    sheet = workbook['鉴权终端用户数据']

    # 遍历行
    for row in sheet.iter_rows(min_row=2):
        # 获取单元格数据
        real_name = row[0].value
        msisdn = row[1].value
        id_card = row[2].value
        group_name = row[3].value
        department = row[4].value
        auth_type = row[5].value
        auth_name = row[6].value
        auth_pass = row[7].value
        mac_addr = row[10].value
        userValues_BIZ.append((1, auth_name, auth_pass, auth_type, created_time, department, group_name, id_card, '', 0, 0, mac_addr, msisdn, '', created_time, created_time, real_name, '', update_time, update_time, 0))
        # print(userValues_BIZ)
        # 如果userValues_BIZ列表长度达到1000，则批量插入到数据库中
        if len(userValues_BIZ) == 1000:
            print("开始了：")
            mysql_BIZ.insert_new_data(sql_BIZ_1, userValues_BIZ)
            userValues_BIZ = []

    # 如果还有未插入的数据，则批量插入到数据库中
    if len(userValues_BIZ) > 0:
        mysql_BIZ.insert_new_data(sql_BIZ_1, userValues_BIZ)

    mysql_BIZ.commit()
def Insert_AuthCPEData_to_mysql(filename):
    # 读取Excel文件
    mysql_BIZ = DB_BIZ()
    mysql_BIZ.init()

    created_time = time.strftime('%Y-%m-%d %H:%M:%S')
    update_time = time.strftime('%Y-%m-%d %H:%M:%S')
    userValues_BIZ = []
    sql_BIZ_1 = 'INSERT INTO db_eiacs_biz.aaa_auth_gateway_device (active_num, auth_name, auth_pass, auth_type, create_time, group_name, id_str, ip, is_online,  mac_addr, metric, msisdn, name, nas_id_str, offline_time, online_time, remark,route_ip_pool,update_time, session_expire_time) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook['鉴权CPE设备数据']

    # 遍历行
    for row in sheet.iter_rows(min_row=2):
        # 获取单元格数据
        id_str = row[0].value
        name = row[1].value
        msisdn = row[2].value
        group_name = row[3].value
        auth_type = row[4].value
        auth_name = row[5].value
        auth_pass = encode_password(row[6].value)
        mac_addr = row[9].value
        remark = row[11].value
        route_ip_pool = row[10].value
        userValues_BIZ.append((1, auth_name, auth_pass, auth_type, created_time, group_name, id_str, '', '', mac_addr, 1, msisdn, name, '',update_time, update_time,remark,route_ip_pool, update_time, update_time))
        # print(userValues_BIZ)
        # 如果userValues_BIZ列表长度达到1000，则批量插入到数据库中
        if len(userValues_BIZ) == 1000:
            print("开始了：")
            mysql_BIZ.insert_new_data(sql_BIZ_1, userValues_BIZ)
            userValues_BIZ = []

    # 如果还有未插入的数据，则批量插入到数据库中
    if len(userValues_BIZ) > 0:
        mysql_BIZ.insert_new_data(sql_BIZ_1, userValues_BIZ)

    mysql_BIZ.commit()

if __name__ == '__main__':
    time_start = time.time()
    print("start time: ", time.strftime('%Y-%m-%d %H:%M:%S'))
    # Insert_sign_user_new('/Users/hejian/Desktop/联通数科/性能测试数据/鉴权性能测试数据/user/auth-user-200000-20230420163033.xlsx')
    time_end = time.time()
    print("finish time: ", time.strftime('%Y-%m-%d %H:%M:%S'))
    print("cost time: ", time_end - time_start)
    # Insert_AuthCPEData_to_mysql('/Users/hejian/Desktop/联通数科/性能测试数据/鉴权性能测试数据/cpe/auth-CPE-200000-20230420175120.xlsx')
    # Insert_AuthData_to_mysql('/Users/hejian/Desktop/联通数科/性能测试数据/鉴权性能测试数据/user/auth-user-200000-20230420163033.xlsx')
    Insert_sign_user_new('/Users/hejian/Desktop/联通数科/性能测试数据/鉴权性能测试数据/user/auth-user-200000-20230420163033.xlsx')
    # Delete_sign_user('/Users/hejian/Desktop/联通数科/性能测试数据/鉴权性能测试数据/auth-user-10000-20230331115920.xlsx')
    # Delete_sign_user_new('/Users/hejian/Desktop/联通数科/性能测试数据/鉴权性能测试数据/auth-user-10000-20230331115920.xlsx')
    # Delete_common_user_new('/Users/hejian/Desktop/联通数科/性能测试数据/签约性能测试数据/sign-50000-20230403153329.xlsx')
    # Delete_common_user_auth_user('/Users/hejian/Desktop/联通数科/性能测试数据/鉴权性能测试数据/user/auth-user-10000-20230403155805.xlsx')
    time_end = time.time()
    print("finish time: ", time.strftime('%Y-%m-%d %H:%M:%S'))
    print("cost time: ", time_end - time_start)

