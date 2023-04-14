import csv
import random
import string
import time
import uuid
from random import randint
import openpyxl
import datetime
import pandas as pd
from id_num import generate_id_number

now = datetime.datetime.now()
formatted_time = now.strftime('%Y%m%d%H%M%S')

IP_SD= ["源IP","目的IP"]
# IP_group = ["192.168.0.1/24","192.168.0.1/16","192.168.0.1/32","192.168.0.1/8"]

create_at = time.strftime('%Y-%m-%d %H:%M:%S')
update_at = time.strftime('%Y-%m-%d %H:%M:%S')

str_phone = list()
str_id = list()
str_id_connt = list()
str_phone_connt = list()


def generate_ip():
    # 随机生成4个数字作为IP地址的四个部分
    ip_parts = [str(random.randint(0, 255)) for _ in range(4)]
    # 随机生成一个掩码长度
    mask_len = random.randint(8, 24)
    # 将四个数字拼接成IP地址
    ip_address = '.'.join(ip_parts)
    # 将IP地址和掩码长度拼接成IP地址段
    ip_segment = f"{ip_address}/{mask_len}"
    return ip_segment

def generate_domain():
    # 随机生成一个顶级域名
    top_level_domains = ['.com', '.net', '.org', '.gov', '.edu', '.biz']
    top_level_domain = random.choice(top_level_domains)

    # 随机生成一个二级域名
    second_level_domain = ''.join(random.choices(string.ascii_lowercase, k=10))

    # 随机生成一个三级域名
    third_level_domain = ''.join(random.choices(string.ascii_lowercase, k=5))

    # 组合成完整的域名地址
    domain = f"{third_level_domain}.{second_level_domain}{top_level_domain}"

    return domain
def get_quanyu_IP(num):
    wb = openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.title = "转发设备IP地址对象案例"
    sheet2 = wb.create_sheet("转发设备IP地址对象模板")

    # 表头
    titles = ('IP地址对象名称', '类型', 'IPV4地址段', '描述')
    for index, title in enumerate(titles):
        sheet2.cell(row=1, column=index + 1, value=title)
    # 生成数据
    for r in range(2, num + 2):
        IP_radum = generate_ip()
        sheet2.cell(r, 1, value="IP对象"+str(r))
        sheet2.cell(r, 2, value=random.choice(IP_SD))
        sheet2.cell(r, 3, value= IP_radum)
        sheet2.cell(r, 4, value='python')
        global str_phone
        global str_id
        str_phone = list()
        str_id = list()
    # 设置列宽
    for col in ['B', 'C', 'D', 'G', 'J', 'K', 'L']:
        sheet2.column_dimensions[col].width = 15
    sheet2.column_dimensions['L'].width = 20
    # 保存为Excel和CSV文件
    file_name = f'quanyu-IP-{num}-{formatted_time}'
    wb.save(f'/Users/hejian/Desktop/联通数科/性能测试数据/权域性能测试数据/{file_name}.xlsx')
    df = pd.read_excel(f'/Users/hejian/Desktop/联通数科/性能测试数据/权域性能测试数据/{file_name}.xlsx', sheet_name='转发设备IP地址对象模板')
    df.to_csv(f'/Users/hejian/Desktop/联通数科/性能测试数据/权域性能测试数据/{file_name}.csv', index=False)


def get_quanyu_domain(num):
    wb = openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.title = "转发设备域名地址对象案例"
    sheet2 = wb.create_sheet("转发设备域名地址对象模板")

    # 表头
    titles = ('域名地址对象名称', '域名地址', '描述')
    for index, title in enumerate(titles):
        sheet2.cell(row=1, column=index + 1, value=title)
    # 生成数据
    for r in range(2, num + 2):
        IP_radum = generate_ip()
        sheet2.cell(r, 1, value="域名对象" + str(r))
        sheet2.cell(r, 2, value=generate_domain())
        sheet2.cell(r, 4, value='python')
        global str_phone
        global str_id
        str_phone = list()
        str_id = list()
    # 设置列宽
    for col in ['B', 'C', 'D', 'G', 'J', 'K', 'L']:
        sheet2.column_dimensions[col].width = 15
    sheet2.column_dimensions['L'].width = 20
    # 保存为Excel和CSV文件
    file_name = f'quanyu-domain-{num}-{formatted_time}'
    wb.save(f'/Users/hejian/Desktop/联通数科/性能测试数据/权域性能测试数据/{file_name}.xlsx')
    df = pd.read_excel(f'/Users/hejian/Desktop/联通数科/性能测试数据/权域性能测试数据/{file_name}.xlsx',
                       sheet_name='转发设备域名地址对象模板')
    df.to_csv(f'/Users/hejian/Desktop/联通数科/性能测试数据/权域性能测试数据/{file_name}.csv', index=False)


if __name__ == '__main__':
    # get_auth_cpe_user_gpt_2(12)
    get_quanyu_IP(100)
    get_quanyu_domain(100)